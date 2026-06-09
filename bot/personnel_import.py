from __future__ import annotations

from dataclasses import dataclass
from io import BytesIO
from typing import Any

from openpyxl import load_workbook


@dataclass(frozen=True)
class PersonnelImportRow:
    row_number: int
    department_name: str
    personnel_name: str
    extension: str | None


def parse_personnel_workbook(content: bytes) -> tuple[list[PersonnelImportRow], list[str]]:
    workbook = load_workbook(BytesIO(content), data_only=True, read_only=True)
    try:
        sheet = workbook.worksheets[0]
        rows: list[PersonnelImportRow] = []
        errors: list[str] = []
        for row_number, row in enumerate(sheet.iter_rows(min_row=2, max_col=3, values_only=True), start=2):
            department_name = _cell_to_text(row[0])
            personnel_name = _cell_to_text(row[1])
            extension = _cell_to_text(row[2])
            if not department_name and not personnel_name and not extension:
                continue
            if not department_name:
                errors.append(f"Satır {row_number}: departman boş")
                continue
            if not personnel_name:
                errors.append(f"Satır {row_number}: personel adı boş")
                continue
            rows.append(
                PersonnelImportRow(
                    row_number=row_number,
                    department_name=department_name,
                    personnel_name=personnel_name,
                    extension=extension or None,
                )
            )
        return rows, errors
    finally:
        workbook.close()


def _cell_to_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, str):
        return value.strip()
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()